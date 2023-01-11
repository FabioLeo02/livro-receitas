#FEITO POR FABIO LEANDRO LOPES DA CUNHA
#CODIGO PARA JUNÇÃO DE DADOS DE PLANILHAS DIFERENTES EM UMA SÓ

#import OPENPYXL - Importação da biblioteca openpyxl
import openpyxl

#Abre a planilha deseja com os dados     
wb = openpyxl.load_workbook("C:/Users/A2248/Documents/Downloads/RELATORIO TESTE.xlsx") 
ws = wb.active

#Salva os dados da cedula escolhida em variaveis 
Ems = ws['E2'].value
Emsres = ws['G2'].value
Sub = ws['B5'].value
Subres = ws['C5'].value
Ali = ws['D5'].value
Alires = ws['E5'].value
Comp = ws['B6'].value
Compres = ws['C6'].value
Pcf = ws['B32'].value
Pcfres = ws['E32'].value
Pcn = ws['B33'].value
Pcnres = ws['E33'].value
DataImp = ws['B253'].value
DataImpres = ws['D253'].value

#Abre nova 
novatab = openpyxl.Workbook()
nvs = novatab.active

#coloca os valores de uma planilha pedida na nova planilha

dados = [
      [Ems,Emsres],
      [Sub,Subres],
      [Ali,Alires],
      [Comp,Compres],
      [Pcf,Pcfres],
      [Pcn,Pcnres],
      [DataImp,DataImpres],
]

for i in dados:
    nvs.append(i)

nvs['B1'].number_format = 'dd/mm/yyyy'
nvs['B7'].number_format = 'dd/mm/yyyy'


#salva os valores na nova planilha
novatab.save("C:/Users/A2248/Documents/Downloads/Teste.xlsx")

